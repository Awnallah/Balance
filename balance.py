from openpyxl import load_workbook, Workbook


def createDic(file, worksheet):
	wb = load_workbook(filename=file)
	sheet1 = wb[worksheet]
	dic = {}
	for row in sheet1.rows[1:]:
		key = row[0].value + ' ' + row[1].value + ' ' + row[4].value
		val = -row[7].value - row[8].value - row[9].value
		dic[key]= val

	return dic



july_dic = createDic('FAMIS_July.xlsx', 'CWWSIP')
august_dic = createDic('FAMIS_August.xlsx', 'CWWSIP')


def balance(new_book, sheet, august_dic, july_dic, output):
	#new_book = Workbook()
	new_sheet = new_book.create_sheet(sheet)
	new_sheet.append(['TEN','PROJECT', 'INDEX_CODE', 'August balance', 'July balance', 'Delta'])
	for key in august_dic:
		if key in july_dic:
			bal = august_dic[key] - july_dic[key]
			keys = key.split()
			new_sheet.append(keys + [august_dic[key], july_dic[key], bal])
	

#balance('balance.xlsx')

def init(file1='FAMIS_July.xlsx', file2='FAMIS_August.xlsx', output='balanceOut.xlsx'):
	wb = load_workbook(file1)
	sheets = wb.get_sheet_names()
	new_book = Workbook()
	for sheet in sheets:
		print sheet
		july_dic = createDic(file1, sheet)
		august_dic = createDic(file2, sheet)
		balance(new_book ,sheet, august_dic, july_dic, output)

	new_book.save(output)



init()

